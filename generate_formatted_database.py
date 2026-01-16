#!/usr/bin/env python3
"""
Generate a formatted Excel file from the pickup database CSV
with color coding and borders for easy visualization
"""

import csv
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("❌ openpyxl not installed. Installing...")
    import subprocess
    import sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Wire color hex codes
WIRE_COLORS = {
    'Black': '000000',
    'White': 'FFFFFF',
    'Red': 'DC2626',
    'Green': '16A34A',
    'Blue': '2563EB',
    'Yellow': 'EAB308',
    'Brown': '92400E',
    'Orange': 'EA580C',
    'Purple': '9333EA',
    'Gray': '6B7280',
    'Bare/Shield': 'D4D4D4'
}

def get_text_color(bg_color):
    """Determine if text should be black or white based on background"""
    # White background colors need black text
    light_colors = ['FFFFFF', 'EAB308', 'D4D4D4', 'EA580C']
    return '000000' if bg_color in light_colors else 'FFFFFF'

def create_formatted_excel(csv_file='pickup_database_detailed.csv', output_file='pickup_database_formatted.xlsx'):
    """Create a formatted Excel file with colors and borders"""

    print(f"📊 Creating formatted Excel file from {csv_file}...")

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Pickup Database"

    # Define styles
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    # Hot/Ground section fills
    standard_section_fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
    reversed_section_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    # Borders
    thick_border = Border(
        left=Side(style='thick', color='000000'),
        right=Side(style='thick', color='000000'),
        top=Side(style='thick', color='000000'),
        bottom=Side(style='thick', color='000000')
    )

    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # Read CSV
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)

    # Column indices for special formatting
    STANDARD_GROUND_COL = 11  # Column K (0-indexed: 10)
    STANDARD_HOT_COL = 12     # Column L
    REVERSED_GROUND_COL = 13  # Column M
    REVERSED_HOT_COL = 14     # Column N

    # Write header row
    header_row = rows[0]
    for col_idx, header in enumerate(header_row, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

        # Make hot/ground columns stand out more
        if col_idx in [STANDARD_GROUND_COL, STANDARD_HOT_COL, REVERSED_GROUND_COL, REVERSED_HOT_COL]:
            cell.border = thick_border

    # Write data rows
    for row_idx, row in enumerate(rows[1:], start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border

            # Color code wire color cells
            if col_idx in [3, 4, 6, 7, 9, 10, 11, 12, 13, 14]:  # Wire color columns
                if value in WIRE_COLORS:
                    bg_color = WIRE_COLORS[value]
                    text_color = get_text_color(bg_color)
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                    cell.font = Font(bold=True, color=text_color, size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # Special formatting for HOT/GROUND columns
            if col_idx == STANDARD_GROUND_COL:  # STANDARD Ground
                cell.fill = standard_section_fill
                cell.border = thick_border
                if value in WIRE_COLORS:
                    cell.font = Font(bold=True, color='16A34A', size=11)  # Green for ground

            elif col_idx == STANDARD_HOT_COL:  # STANDARD Hot
                cell.fill = standard_section_fill
                cell.border = thick_border
                if value in WIRE_COLORS:
                    cell.font = Font(bold=True, color='DC2626', size=11)  # Red for hot

            elif col_idx == REVERSED_GROUND_COL:  # PHASE REVERSED Ground
                cell.fill = reversed_section_fill
                cell.border = thick_border
                if value in WIRE_COLORS:
                    cell.font = Font(bold=True, color='16A34A', size=11)  # Green for ground

            elif col_idx == REVERSED_HOT_COL:  # PHASE REVERSED Hot
                cell.fill = reversed_section_fill
                cell.border = thick_border
                if value in WIRE_COLORS:
                    cell.font = Font(bold=True, color='DC2626', size=11)  # Red for hot

    # Set column widths
    column_widths = {
        'A': 25,  # Preset Name
        'B': 20,  # Manufacturer
        'C': 20,  # North RED
        'D': 22,  # North BLACK
        'E': 15,  # North Pole
        'F': 20,  # South RED
        'G': 22,  # South BLACK
        'H': 15,  # South Pole
        'I': 22,  # Series 1
        'J': 22,  # Series 2
        'K': 22,  # STD Ground
        'L': 18,  # STD Hot
        'M': 22,  # REV Ground
        'N': 18,  # REV Hot
        'O': 35   # Notes
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Set row height for header
    ws.row_dimensions[1].height = 40

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Save workbook
    wb.save(output_file)
    print(f"✅ Created {output_file}")
    print(f"\n📝 Formatting applied:")
    print(f"   • Wire colors are color-coded with matching backgrounds")
    print(f"   • STANDARD Ground/Hot columns have BLUE background")
    print(f"   • PHASE REVERSED Ground/Hot columns have YELLOW background")
    print(f"   • Ground wires are GREEN text")
    print(f"   • Hot wires are RED text")
    print(f"   • Thick borders around hot/ground columns")
    print(f"\n🎨 Color legend:")
    for color, hex_code in WIRE_COLORS.items():
        print(f"   • {color}: #{hex_code}")

if __name__ == '__main__':
    create_formatted_excel()
